"""
Excel文件处理服务
"""
import openpyxl
from pathlib import Path
from typing import List, Dict, Any
from loguru import logger


class ExcelService:
    """Excel文件处理服务"""
    
    @staticmethod
    def split_excel_file(
        source_file_path: str,
        output_dir: Path,
        batch_session_id: int
    ) -> List[Dict[str, Any]]:
        """
        拆分多Sheet Excel文件为多个单Sheet文件
        
        Args:
            source_file_path: 源文件路径
            output_dir: 输出目录
            batch_session_id: 批量会话ID
            
        Returns:
            List[Dict]: 每个Sheet的信息列表，包含：
                - sheet_name: Sheet名称
                - sheet_index: Sheet索引（从0开始）
                - split_file_path: 拆分后的文件路径
        """
        # 创建输出目录
        output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"[ExcelService] 开始拆分Excel文件 - source={source_file_path}, output_dir={output_dir}")
        
        # 读取源文件
        try:
            source_workbook = openpyxl.load_workbook(source_file_path, read_only=False)
            sheet_names = source_workbook.sheetnames
            logger.info(f"[ExcelService] 检测到 {len(sheet_names)} 个Sheet: {sheet_names}")
        except Exception as e:
            logger.error(f"[ExcelService] 读取Excel文件失败: {str(e)}")
            raise Exception(f"读取Excel文件失败: {str(e)}")
        
        split_files = []
        
        try:
            for index, sheet_name in enumerate(sheet_names):
                logger.info(f"[ExcelService] 处理Sheet {index + 1}/{len(sheet_names)}: {sheet_name}")
                
                # 创建新的工作簿，只包含当前Sheet
                new_workbook = openpyxl.Workbook()
                new_sheet = new_workbook.active
                new_sheet.title = sheet_name
                
                # 复制数据
                source_sheet = source_workbook[sheet_name]
                
                # 复制所有行数据
                for row in source_sheet.iter_rows(values_only=True):
                    new_sheet.append(row)
                
                # 复制列宽（如果存在）
                if hasattr(source_sheet, 'column_dimensions'):
                    for col_letter, dimension in source_sheet.column_dimensions.items():
                        if dimension.width:
                            new_sheet.column_dimensions[col_letter].width = dimension.width
                
                # 保存为独立文件
                # 清理Sheet名称中的特殊字符，确保文件名安全
                safe_sheet_name = "".join(
                    c for c in sheet_name 
                    if c.isalnum() or c in (' ', '-', '_')
                ).strip()
                
                # 如果清理后为空，使用默认名称
                if not safe_sheet_name:
                    safe_sheet_name = f"Sheet{index}"
                
                output_filename = f"sheet_{index}_{safe_sheet_name}.xlsx"
                output_path = output_dir / output_filename
                
                # 保存文件
                new_workbook.save(output_path)
                new_workbook.close()
                
                # 验证文件是否真的被创建
                if not output_path.exists():
                    raise Exception(f"文件保存失败，文件不存在: {output_path}")
                
                file_size = output_path.stat().st_size
                logger.info(f"[ExcelService] 已拆分Sheet: {sheet_name} -> {output_path} (大小: {file_size} bytes)")
                
                # 使用 as_posix() 确保路径使用正斜杠，跨平台兼容
                split_files.append({
                    "sheet_name": sheet_name,
                    "sheet_index": index,
                    "split_file_path": output_path.as_posix()  # 使用正斜杠，跨平台兼容
                })
        
        finally:
            # 确保关闭源工作簿
            source_workbook.close()
        
        logger.info(f"[ExcelService] 拆分完成 - 共生成 {len(split_files)} 个文件")
        return split_files
    
    @staticmethod
    def get_sheet_count(file_path: str) -> int:
        """
        获取Excel文件的Sheet数量
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            int: Sheet数量
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            count = len(workbook.sheetnames)
            workbook.close()
            return count
        except Exception as e:
            logger.error(f"[ExcelService] 获取Sheet数量失败: {str(e)}")
            raise Exception(f"获取Sheet数量失败: {str(e)}")

